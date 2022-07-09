def get_data_fpbase(root):
    from selenium import webdriver
    from fake_useragent import UserAgent
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    from Bio import SeqIO
    from Bio.Seq import Seq
    from Bio.SeqRecord import SeqRecord

    from openpyxl import Workbook
    from openpyxl.styles.fonts import Font

    import math

    ua = UserAgent()
    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent={ua.random}')
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    leaf_list = get_tree_fpbase(driver, root)

    ans_list = []
    seq_list = []
    for name, link in leaf_list:
        name = name.replace(' ', '-')
        table_list = crawl_dataset(link)
        if len(table_list) > 1:
            raw_list = table_list[1].values.tolist()
            seq, ref = crawl_data(driver, link)
            if len(raw_list) > 1:
                ans_list.append([name, *min(raw_list, key=lambda t: t.count(math.nan))[1:], ref])
            else:
                ans_list.append([name, *raw_list[0], ref])

            seq_list.append(SeqRecord(seq=Seq(seq), id=name.replace(' ', '-'), name="", description=""))
            print(name, ':', len(seq))
        else:
            print(name)

    driver.quit()

    print("Length : %d" % len(ans_list))
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col, val in enumerate(["Name", "Ex λ", "Em λ", "EC (/M * cm)", "QY", "Brightness", "pKa", "Maturation", "Lifetime", "Reference"]):
        ws.cell(row=1, column=col + 1, value=val).font = Font(bold=True)

    for row, clist in enumerate(ans_list):
        for col, val in enumerate(clist):
            if math.nan != val:
                ws.cell(row=row + 2, column=col + 1, value=val)

    wb.save("Data/%s_data.xlsx" % root)

    SeqIO.write(seq_list, "Data/%s_data.fasta" % root, "fasta")


def get_tree_fpbase(driver, root):
    from time import sleep
    driver.get("https://www.fpbase.org/protein/" + root)
    sleep(1)

    return driver.execute_script(
        "return [...document.querySelectorAll('.lineage-wrapper svg')[0].getElementsByClassName('node')].map(v=>["
        "v.textContent, v.getElementsByTagName('a')[0].href.baseVal])")


def crawl_dataset(tar):
    import pandas as pd
    return pd.read_html("https://www.fpbase.org" + tar)


def crawl_data(driver, tar):
    from selenium.webdriver.common.by import By
    from time import sleep
    driver.get("https://www.fpbase.org" + tar)
    sleep(0.5)

    return driver.execute_script("return document.querySelector('.formatted_aminosquence').textContent."
                                 "replaceAll(' ','')"), driver.find_element(By.CLASS_NAME, "reference-details").find_element(By.CSS_SELECTOR, "a").get_attribute("href")
